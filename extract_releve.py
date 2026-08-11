"""Extrait les transactions d'un releve de compte bancaire (PDF) vers Excel.

Le PDF ne contient aucune couche texte : chaque caractere y est dessine sous
forme de contours vectoriels. Le texte est donc reconstruit en rasterisant
chaque cellule de la grille monospace et en la comparant aux empreintes de
glyphes apprises (glyph_model.npz).

Usage:
    python extract_releve.py releve.pdf [-o releve.xlsx]
"""
from __future__ import annotations

import argparse
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import numpy as np
import pdfplumber
from PIL import Image, ImageDraw

MODEL = Path(__file__).with_name('glyph_model.npz')

# --- geometrie de la page -------------------------------------------------
TOP_CUT = 60.0        # ignore l'en-tete de la banque (micro-police + logo)
ANCHOR_MIN_H = 4.5    # hauteur mini d'un trace pour servir d'ancre de ligne
SS = 8                # suréchantillonnage (px par point)
CELL_UP, CELL_DOWN = 8.6, 2.8
GRID = (12, 18)
MAX_DIST = 0.55       # distance max acceptee entre un glyphe et son modele

COL_RECORD = 14       # colonne des lignes "solde" et en-tete de transaction
COL_DETAIL = 16       # colonne des lignes de detail


# =========================================================================
# 1. Reconstruction du texte
# =========================================================================
def _components(page):
    out = []
    for o in page.curves + page.rects:
        if not o.get('fill') or o['bottom'] < TOP_CUT:
            continue
        pts = o.get('pts') or [(o['x0'], o['top']), (o['x1'], o['top']),
                               (o['x1'], o['bottom']), (o['x0'], o['bottom'])]
        out.append({'x0': o['x0'], 'top': o['top'], 'x1': o['x1'],
                    'bottom': o['bottom'], 'pts': [tuple(p) for p in pts]})
    return out


def _baselines(comps, tol=1.5, min_gap=5.0):
    ys = sorted(c['bottom'] for c in comps if c['bottom'] - c['top'] >= ANCHOR_MIN_H)
    cand, cur = [], []
    for y in ys:
        if cur and y - cur[0] > tol:
            cand.append((sum(cur) / len(cur), len(cur)))
            cur = []
        cur.append(y)
    if cur:
        cand.append((sum(cur) / len(cur), len(cur)))

    # les parentheses et jambages descendent sous la ligne et creent de
    # fausses lignes : on les replie sur la ligne reelle la mieux etayee
    merged = []
    for y, n in cand:
        if merged and y - merged[-1][0] < min_gap:
            if n > merged[-1][1]:
                merged[-1] = (y, n)
        else:
            merged.append((y, n))
    return [y for y, _ in merged]


def _lines(comps):
    bl = _baselines(comps)
    buckets = [[] for _ in bl]
    for c in comps:
        best, bd = None, 1e9
        for i, y in enumerate(bl):
            d = c['bottom'] - y
            if -9.0 <= d <= 2.5 and abs(d) < bd:
                best, bd = i, abs(d)
        if best is not None:
            buckets[best].append(c)
    return [(bl[i], b) for i, b in enumerate(buckets) if b]


def _fit_grid(all_x0):
    """Retrouve le pas et l'origine de la grille monospace."""
    x = np.asarray(all_x0)
    pitch = max(np.arange(5.80, 6.20, 0.0005),
                key=lambda p: abs(np.mean(np.exp(1j * (x % p) / p * 2 * np.pi))))
    m = np.mean(np.exp(1j * (x % pitch) / pitch * 2 * np.pi))
    return pitch, (np.angle(m) % (2 * np.pi)) / (2 * np.pi) * pitch


def _raster(comps, cell_x, baseline, pitch):
    w = int(round(pitch * SS)) + 4
    h = int(round((CELL_UP + CELL_DOWN) * SS))
    x0, y0 = cell_x - 2.0 / SS, baseline - CELL_UP
    acc = np.zeros((h, w), dtype=bool)
    for c in comps:
        layer = Image.new('1', (w, h), 0)
        poly = [((x - x0) * SS, (y - y0) * SS) for x, y in c['pts']]
        draw = ImageDraw.Draw(layer)
        if len(poly) >= 3:
            draw.polygon(poly, fill=1)          # remplissage pair-impair :
        else:                                   # les contre-formes (o, e, 8)
            draw.rectangle(poly, fill=1)        # s'obtiennent par XOR
        acc ^= np.asarray(layer, dtype=bool)
    img = Image.fromarray((acc * 255).astype(np.uint8))
    v = np.asarray(img.resize(GRID, Image.BOX), dtype=np.float32).ravel() / 255.0
    return np.concatenate([v / (np.linalg.norm(v) + 1e-6), [np.sqrt(v.sum())]])


def read_text(pdf_path):
    """-> [(page, colonne de depart, texte)] pour chaque ligne du releve."""
    model = np.load(MODEL, allow_pickle=True)
    centroids, chars = model['centroids'], model['chars']

    with pdfplumber.open(pdf_path) as pdf:
        pages = [_components(p) for p in pdf.pages]
    abscisses = [c['x0'] for pg in pages for c in pg]
    if not abscisses:
        raise ValueError(
            f'{pdf_path} : aucun trace vectoriel exploitable. Soit ce PDF n\'est '
            'pas un releve de ce format, soit il porte deja une couche texte -- dans '
            'ce cas, un extracteur de texte ordinaire suffit.')
    pitch, origin = _fit_grid(abscisses)

    unknown = 0
    out = []
    for pno, comps in enumerate(pages, start=1):
        for base, line in _lines(comps):
            cells = {}
            for c in line:
                cells.setdefault(int(round((c['x0'] - origin) / pitch)), []).append(c)
            decoded = {}
            for col, cs in cells.items():
                f = _raster(cs, origin + col * pitch, base, pitch)
                d = np.linalg.norm(centroids - f, axis=1)
                j = int(d.argmin())
                if d[j] > MAX_DIST:
                    unknown += 1
                decoded[col] = str(chars[j]) if d[j] <= MAX_DIST else '?'
            start = min(decoded)
            text = ''.join(decoded.get(i, ' ') for i in range(start, max(decoded) + 1))
            out.append((pno, start, text.rstrip()))
    if unknown:
        print(f'Attention : {unknown} caractere(s) non reconnu(s), notes "?"')
    return out


# =========================================================================
# 2. Analyse du releve
# =========================================================================
# En gras, le glyphe du chiffre 0 et celui de la lettre O sont identiques :
# dans les champs numeriques on force la lecture en chiffre.
def _digits(s):
    return s.replace('O', '0')


def _amount(s):
    return float(_digits(s).replace('.', '').replace(',', '.'))


def _date(s):
    return datetime.strptime(_digits(s), '%d-%m-%Y').date()


RE_SOLDE = re.compile(
    r'^S[O0]LDE AU\s+([O0-9]{2}-[O0-9]{2}-[O0-9]{4})(?:\s+[O0-9]{2}:[O0-9]{2})?'
    r'\s+([A-Z]{3})\s+([+-])\s*([O0-9.]*[O0-9],[O0-9]{2})$')

RE_HEADER = re.compile(
    r'^([0-9]{4})\s+([0-9]{2}-[0-9]{2}-[0-9]{4})\s+\(VAL\.\s*([0-9]{2}-[0-9]{2}-[0-9]{4})\)'
    r'\s+([+-])\s*([0-9.]*[0-9],[0-9]{2})$')

RE_IBAN = re.compile(r'\b([A-Z]{2}[0-9]{2}(?:\s[0-9]{4}){2,4})\b')
RE_STRUCT = re.compile(r'\+\+\+[0-9]{3}/[0-9]{4}/[0-9]{5}\+\+\+')
RE_CARTE = re.compile(r'CARTE\s+N?[°\s]*([0-9X]{4}(?:\s+[0-9X]{4}){3})')
RE_TIERS = re.compile(r'\b(?:VERS|POUR|DE|DU)\s+(?:[A-Z]{2}[0-9]{2}(?:\s[0-9]{4}){2,4}\s+)?')
RE_FIN_TIERS = re.compile(
    r'\s+(?:COMMUNICATION|REFERENCE|REF\.|VERS|EFFECTUE|/A/|\+\+\+|[0-9]{6,})')

TYPES = [
    ('ACHAT OPC', 'Achat OPC'),
    ('VOTRE DOMICILIATION', 'Domiciliation'),
    ('ORDRE PERMANENT', 'Ordre permanent'),
    ('VIREMENT INSTANTANE', 'Virement instantane'),
    ('VIREMENT VIA WERO', 'Virement Wero'),
    ('VIREMENT', 'Virement'),
    ('VERSEMENT INSTANTANE', 'Versement instantane recu'),
    ('VERSEMENT', 'Versement recu'),
    ('BANCONTACT', 'Paiement Bancontact'),
    ('APP BANCONTACT', 'Paiement mobile P2P'),
    ('PAIEMENT DEBITMASTERCARD', 'Paiement Mastercard'),
    ('RELEVE MASTERCARD', 'Releve Mastercard'),
    ('PARTICIPATION AUX FRAIS', 'Frais de gestion'),
    ('RETRAIT', 'Retrait'),
]


@dataclass
class Operation:
    numero: str
    date_operation: date
    date_valeur: date
    montant: float
    page: int
    lignes: list = field(default_factory=list)
    compte: str = ''            # IBAN du titulaire, pour ecarter son propre compte

    @property
    def description(self):
        texte = ''
        for ligne in self.lignes:
            # un mot coupe en fin de ligne se termine par un trait d'union colle
            colle = texte.endswith('-') and not texte.endswith(' -')
            texte += ligne if (colle or not texte) else ' ' + ligne
        return texte

    @property
    def sens(self):
        return 'Credit' if self.montant > 0 else 'Debit'

    @property
    def type(self):
        d = self.description
        for motif, libelle in TYPES:
            if d.startswith(motif):
                return libelle
        return 'Autre'

    @property
    def iban_contrepartie(self):
        for m in RE_IBAN.finditer(self.description):
            if m.group(1).replace(' ', '') != self.compte:
                return m.group(1)
        return ''

    @property
    def contrepartie(self):
        d = self.description
        if self.type == 'Paiement Bancontact':
            parts = [p.strip() for p in d.split(' - ')]
            return parts[2] if len(parts) > 2 else ''
        if self.type == 'Paiement Mastercard':
            m = re.search(r'\d{2}/\d{2}\s+(.*?)\s+[\d.,]+\s+EUR', d)
            return m.group(1) if m else ''
        m = RE_TIERS.search(d)
        if not m:
            return ''
        reste = d[m.end():]
        coupe = RE_FIN_TIERS.search(reste)
        return reste[:coupe.start()].strip() if coupe else reste.strip()

    @property
    def communication(self):
        m = re.search(r'COMMUNICATION\s*:\s*(.*?)(?:\s+REFERENCE DU CREANCIER|$)',
                      self.description)
        if m:
            return m.group(1).strip()
        m = RE_STRUCT.search(self.description)
        return m.group(0) if m else ''

    @property
    def reference(self):
        m = re.search(r'(?:REFERENCE DU CREANCIER\s*:|REF\.\s*:?|N° ORDRE)\s*(\S+)',
                      self.description)
        return m.group(1) if m else ''

    @property
    def carte(self):
        m = RE_CARTE.search(self.description)
        return m.group(1) if m else ''


def parse(lines):
    soldes, operations = [], []
    compte = ''          # IBAN du titulaire, lu sur la ligne de separation
    courant = None
    for page, col, text in lines:
        if col == COL_RECORD:
            m = RE_IBAN.search(text)
            if m and not compte and text.startswith('-'):
                compte = m.group(1).replace(' ', '')

            m = RE_SOLDE.match(text)
            if m:
                courant = None
                jour, devise, signe, montant = m.groups()
                soldes.append({'date': _date(jour), 'devise': devise,
                               'montant': _amount(montant) * (1 if signe == '+' else -1)})
                continue

            m = RE_HEADER.match(_digits(text))
            if m:
                numero, dop, dval, signe, montant = m.groups()
                courant = Operation(numero=numero, date_operation=_date(dop),
                                    date_valeur=_date(dval), page=page,
                                    montant=_amount(montant) * (1 if signe == '+' else -1))
                operations.append(courant)
                continue
            courant = None
        elif col == COL_DETAIL and courant is not None:
            courant.lignes.append(text.strip())
    for op in operations:
        op.compte = compte
    return soldes, operations


# =========================================================================
# 3. Export Excel
# =========================================================================
COLONNES = ['Date', 'Date valeur', 'N° operation', 'Type', 'Contrepartie',
            'IBAN contrepartie', 'Communication', 'Reference', 'Carte',
            'Sens', 'Debit', 'Credit', 'Montant', 'Solde', 'Description', 'Page']


def to_excel(soldes, operations, chemin):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Releve'
    ws.append(COLONNES)

    solde = soldes[0]['montant']
    lignes = [['', '', '', 'SOLDE INITIAL', '', '', '', '', '', '', None, None,
               None, solde, f"Solde au {soldes[0]['date']:%d/%m/%Y}", '']]
    for op in operations:
        solde += op.montant
        lignes.append([
            op.date_operation, op.date_valeur, op.numero, op.type, op.contrepartie,
            op.iban_contrepartie, op.communication, op.reference, op.carte, op.sens,
            -op.montant if op.montant < 0 else None,
            op.montant if op.montant > 0 else None,
            op.montant, round(solde, 2), op.description, op.page])
    fin = soldes[-1]
    lignes.append(['', '', '', 'SOLDE FINAL', '', '', '', '', '', '', None, None,
                   None, fin['montant'], f"Solde au {fin['date']:%d/%m/%Y}", ''])
    for ligne in lignes:
        ws.append(ligne)

    entete = PatternFill('solid', fgColor='1F3864')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = entete
        cell.alignment = Alignment(vertical='center')
    bordure = PatternFill('solid', fgColor='DDEBF7')
    for ligne in (2, ws.max_row):
        for cell in ws[ligne]:
            cell.font = Font(bold=True)
            cell.fill = bordure

    for lettre, largeur in zip('ABCDEFGHIJKLMNOP',
                               [11, 11, 12, 22, 30, 22, 26, 20, 22,
                                8, 12, 12, 12, 12, 90, 6]):
        ws.column_dimensions[lettre].width = largeur
    for ligne in ws.iter_rows(min_row=2):
        for cell in ligne[:2]:
            cell.number_format = 'DD/MM/YYYY'
        for cell in ligne[10:14]:
            cell.number_format = '#,##0.00'
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLONNES))}{ws.max_row}'

    wb.save(chemin)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('pdf', type=Path, help='releve PDF a depouiller')
    ap.add_argument('-o', '--out', type=Path,
                    help='fichier Excel de sortie (defaut : a cote du script)')
    args = ap.parse_args()
    sortie = args.out or Path(__file__).parent / f'{args.pdf.stem}.xlsx'

    try:
        soldes, operations = parse(read_text(args.pdf))
    except (ValueError, FileNotFoundError) as erreur:
        raise SystemExit(str(erreur))
    if len(soldes) < 2:
        raise SystemExit('Lignes SOLDE introuvables : format de releve inattendu.')

    attendu = soldes[0]['montant'] + sum(op.montant for op in operations)
    ecart = round(attendu - soldes[-1]['montant'], 2)
    print(f'{len(operations)} operations, solde initial {soldes[0]["montant"]:,.2f} EUR, '
          f'solde final {soldes[-1]["montant"]:,.2f} EUR')
    print('Controle : ' + ('OK' if ecart == 0 else f'ECART DE {ecart:.2f} EUR'))

    to_excel(soldes, operations, sortie)
    print(f'Ecrit -> {sortie}')


if __name__ == '__main__':
    main()
