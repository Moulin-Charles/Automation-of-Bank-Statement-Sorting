"""Pre-remplit l'onglet << Saisie depenses >> de Budget_Mensuel.xlsx
a partir d'un releve bancaire extrait par extract_releve.py.

Chaque transaction du releve est passee dans les regles de
regles_categories.json, qui lui donnent une description courte et une
categorie du budget. Les transferts internes et le salaire sont ecartes,
les lignes deja presentes dans le classeur sont detectees et ignorees.

Par defaut le script n'ecrit rien : il affiche le rapport de ce qu'il ferait.
Ajouter --ecrire pour appliquer (une sauvegarde datee est creee avant).

Usage:
    python preremplir_budget.py releve.xlsx
    python preremplir_budget.py releve.xlsx --depuis 2026-06-01 --ecrire
"""
from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from copy import copy

# Emplacement du classeur budget : variable d'environnement si elle existe,
# sinon Documents/Personnel/Budget_Mensuel.xlsx dans le profil utilisateur.
BUDGET_DEFAUT = Path(os.environ.get(
    'BUDGET_MENSUEL', Path.home() / 'Documents' / 'Personnel' / 'Budget_Mensuel.xlsx'))

# Les regles personnelles (regles_categories.local.json) restent hors du depot ;
# a defaut, on retombe sur le jeu d'exemple versionne.
REGLES_LOCALES = Path(__file__).with_name('regles_categories.local.json')
REGLES_EXEMPLE = Path(__file__).with_name('regles_categories.json')
REGLES = REGLES_LOCALES if REGLES_LOCALES.exists() else REGLES_EXEMPLE

FEUILLE = 'Saisie dépenses'
LIGNE1 = 6                 # premiere ligne de donnees
COL = {'date': 2, 'description': 3, 'categorie': 4, 'montant': 5, 'mois': 6}
FIN_RECAP = 205            # borne des SUMIF du recap lateral, a etendre
FIN_VOULUE = 1005          # borne utilisee par les autres onglets

A_CLASSER = 'À CLASSER'

# tolerances de detection des doublons
JOURS_EXACT = 5            # meme montant a +/- 5 jours -> deja saisi
JOURS_PROCHE, ECART_PROCHE = 3, 1.00   # montant voisin -> doublon probable


# =========================================================================
# 1. Regles
# =========================================================================
def _norm(txt: str) -> str:
    """Majuscules sans accent, espaces normalises : cible des motifs."""
    txt = unicodedata.normalize('NFKD', txt)
    txt = ''.join(c for c in txt if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', txt).upper().strip()


@dataclass
class Regle:
    motif: re.Pattern
    description: str = ''
    categorie: str = ''
    raison: str = ''
    ignorer: bool = False
    agreger: bool = False
    verifier: bool = False
    large: bool = False
    montants: dict = field(default_factory=dict)

    def resoudre(self, montant: float) -> tuple[str, str, bool]:
        """-> (description, categorie, a_verifier) pour ce montant."""
        if self.montants:
            cle = f'{abs(montant):.2f}'
            spec = self.montants.get(cle) or self.montants.get('defaut')
            if spec:
                return (spec.get('description', self.description),
                        spec.get('categorie', self.categorie),
                        bool(spec.get('verifier', self.verifier)))
        return self.description, self.categorie, self.verifier


def charger_regles(chemin: Path) -> list[Regle]:
    brut = json.loads(chemin.read_text(encoding='utf-8'))
    regles = []
    for r in brut['regles']:
        regles.append(Regle(
            motif=re.compile(r['motif']),
            description=r.get('description', ''),
            categorie=r.get('categorie', ''),
            raison=r.get('raison', ''),
            ignorer=bool(r.get('ignorer')),
            agreger=bool(r.get('agreger')),
            verifier=bool(r.get('verifier')),
            large=bool(r.get('large')),
            montants=r.get('montants', {}),
        ))
    return regles


# =========================================================================
# 2. Lecture du releve
# =========================================================================
@dataclass
class Operation:
    date: datetime
    montant: float             # positif = depense (le releve compte a l'envers)
    texte: str                 # Type | Contrepartie | Communication, normalise
    texte_large: str           # idem + Description (bruite : ville, n° carte, titulaire)
    contrepartie: str
    apercu: str                # libelle lisible, pour le rapport
    regle: Regle | None = None
    description: str = ''
    categorie: str = ''
    verifier: bool = False
    statut: str = ''           # 'ajout' | 'ignore' | 'doublon' | 'doublon?'
    detail: str = ''


def lire_releve(chemin: Path) -> list[Operation]:
    ws = openpyxl.load_workbook(chemin, data_only=True).worksheets[0]
    lignes = ws.iter_rows(values_only=True)
    entetes = [str(v or '') for v in next(lignes)]
    idx = {h: i for i, h in enumerate(entetes)}
    for requis in ('Date', 'Montant'):
        if requis not in idx:
            raise SystemExit(f'{chemin.name} : colonne "{requis}" absente.')

    ops = []
    for r in lignes:
        date, montant = r[idx['Date']], r[idx['Montant']]
        if not isinstance(date, datetime) or not isinstance(montant, (int, float)):
            continue
        def champ(nom: str) -> str:
            return '' if nom not in idx or r[idx[nom]] is None else str(r[idx[nom]])

        court = ' | '.join(champ(c) for c in ('Type', 'Contrepartie', 'Communication'))
        ops.append(Operation(
            date=date,
            montant=round(-float(montant), 2),
            texte=_norm(court),
            texte_large=_norm(court + ' | ' + champ('Description')),
            contrepartie=champ('Contrepartie'),
            apercu=(champ('Contrepartie') or champ('Type'))[:34],
        ))
    return ops


def classer(ops: list[Operation], regles: list[Regle]) -> None:
    for op in ops:
        for regle in regles:
            if not regle.motif.search(op.texte_large if regle.large else op.texte):
                continue
            op.regle = regle
            if regle.ignorer:
                op.statut, op.detail = 'ignore', regle.raison or 'regle d\'exclusion'
            else:
                op.description, op.categorie, op.verifier = regle.resoudre(op.montant)
                op.description = op.description.replace('{contrepartie}', op.contrepartie)
            break
        else:
            op.description = op.apercu.title()[:28] or 'À identifier'
            op.categorie = A_CLASSER


def agreger(ops: list[Operation]) -> list[Operation]:
    """Fusionne les lignes du meme jour couvertes par une regle 'agreger'."""
    sortie, groupes = [], {}
    for op in ops:
        if op.regle is None or not op.regle.agreger or op.statut == 'ignore':
            sortie.append(op)
            continue
        cle = (op.date, id(op.regle))
        if cle in groupes:
            groupes[cle].montant = round(groupes[cle].montant + op.montant, 2)
            groupes[cle].detail = 'lignes du jour regroupees'
        else:
            groupes[cle] = op
            sortie.append(op)
    return sortie


# =========================================================================
# 3. Classeur budget
# =========================================================================
def lire_saisies(ws) -> list[tuple[datetime, float]]:
    saisies = []
    for ligne in range(LIGNE1, ws.max_row + 1):
        date = ws.cell(ligne, COL['date']).value
        montant = ws.cell(ligne, COL['montant']).value
        if isinstance(date, datetime) and isinstance(montant, (int, float)):
            saisies.append((date, round(float(montant), 2)))
    return saisies


def derniere_ligne(ws) -> int:
    dernier = LIGNE1 - 1
    for ligne in range(LIGNE1, ws.max_row + 1):
        if isinstance(ws.cell(ligne, COL['date']).value, datetime):
            dernier = ligne
    return dernier


def marquer_doublons(ops: list[Operation], saisies: list[tuple[datetime, float]]) -> None:
    """Une saisie existante ne peut absorber qu'une seule operation."""
    libres = list(saisies)
    for op in ops:
        if op.statut == 'ignore':
            continue
        exact = next((s for s in libres
                      if abs(s[1] - op.montant) < 0.005
                      and abs((s[0] - op.date).days) <= JOURS_EXACT), None)
        if exact:
            libres.remove(exact)
            op.statut = 'doublon'
            op.detail = f'deja saisi le {exact[0]:%d/%m}'
            continue
        proche = next((s for s in libres
                       if abs(s[1] - op.montant) <= ECART_PROCHE
                       and abs((s[0] - op.date).days) <= JOURS_PROCHE), None)
        if proche:
            libres.remove(proche)
            op.statut = 'doublon?'
            op.detail = f'proche de {proche[1]:.2f} EUR du {proche[0]:%d/%m}'
            continue
        op.statut = 'ajout'


def categories_valides(wb) -> set[str]:
    ws = wb['Paramètres']
    valides = set()
    for ligne in ws.iter_rows(min_col=6, max_col=6, values_only=True):
        if isinstance(ligne[0], str) and ligne[0].strip():
            valides.add(ligne[0].strip())
    return valides


def etendre_recap(ws) -> bool:
    """Le recap lateral s'arrete a la ligne 205 : l'aligner sur les autres onglets."""
    change = False
    motif = re.compile(rf'(\$[A-Z]{{1,2}}\${LIGNE1}:\$[A-Z]{{1,2}}\$){FIN_RECAP}\b')
    for ligne in ws.iter_rows():
        for cell in ligne:
            v = cell.value
            if isinstance(v, str) and v.startswith('=') and motif.search(v):
                cell.value = motif.sub(rf'\g<1>{FIN_VOULUE}', v)
                change = True
    return change


def ecrire(ws, ops: list[Operation], marquer: bool) -> int:
    depart = derniere_ligne(ws) + 1
    modele = [copy(ws.cell(LIGNE1, c)._style) for c in range(2, 7)]
    for i, op in enumerate(sorted(ops, key=lambda o: o.date)):
        ligne = depart + i
        for j, col in enumerate(range(2, 7)):
            ws.cell(ligne, col)._style = modele[j]
        ws.cell(ligne, COL['date']).value = op.date
        ws.cell(ligne, COL['description']).value = (
            f'{op.description} (?)' if marquer and op.verifier else op.description)
        ws.cell(ligne, COL['categorie']).value = op.categorie
        ws.cell(ligne, COL['montant']).value = op.montant
        ws.cell(ligne, COL['mois']).value = (
            f'=IF(B{ligne}="","",YEAR(B{ligne})&"-"&TEXT(MONTH(B{ligne}),"00"))')
    return depart


# =========================================================================
# 4. Rapport
# =========================================================================
def rapport(ops: list[Operation], categories: set[str]) -> None:
    ajouts = [o for o in ops if o.statut == 'ajout']

    def bloc(titre, lot, detail=False):
        if not lot:
            return
        print(f'\n{titre} ({len(lot)})')
        for o in sorted(lot, key=lambda o: o.date):
            suffixe = f'   <- {o.detail}' if detail and o.detail else ''
            print(f'  {o.date:%d/%m}  {o.montant:8.2f}  {o.description[:26]:26s} '
                  f'{o.categorie[:34]:34s} {o.apercu[:26]:26s}{suffixe}')

    bloc('A AJOUTER', [o for o in ajouts if o.categorie != A_CLASSER and not o.verifier])
    bloc('A AJOUTER - categorie a confirmer', [o for o in ajouts if o.verifier])
    bloc('A AJOUTER - non reconnues, a classer a la main',
         [o for o in ajouts if o.categorie == A_CLASSER and not o.verifier])
    bloc('DOUBLONS PROBABLES - non ajoutes (--forcer pour les ajouter)',
         [o for o in ops if o.statut == 'doublon?'], detail=True)
    bloc('DEJA SAISIES', [o for o in ops if o.statut == 'doublon'], detail=True)
    bloc('IGNOREES', [o for o in ops if o.statut == 'ignore'], detail=True)

    inconnues = {o.categorie for o in ajouts} - categories - {A_CLASSER}
    if inconnues:
        print('\nATTENTION : categories absentes de l\'onglet Parametres -> '
              'elles ne seront comptees nulle part :')
        for c in sorted(inconnues):
            print(f'  - {c}')

    total = sum(o.montant for o in ajouts)
    print(f'\n{len(ajouts)} ligne(s) a ajouter, total {total:.2f} EUR '
          f'({sum(1 for o in ajouts if o.categorie == A_CLASSER)} a classer, '
          f'{sum(1 for o in ajouts if o.verifier)} a confirmer)')


# =========================================================================
def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('releve', type=Path, help='classeur produit par extract_releve.py')
    ap.add_argument('--budget', type=Path, default=BUDGET_DEFAUT)
    ap.add_argument('--regles', type=Path, default=REGLES)
    ap.add_argument('--depuis', help='ne traiter qu\'a partir de cette date (AAAA-MM-JJ)')
    ap.add_argument('--jusqua', help='ne traiter que jusqu\'a cette date (AAAA-MM-JJ)')
    ap.add_argument('--ecrire', action='store_true', help='appliquer les ajouts au classeur')
    ap.add_argument('--forcer', action='store_true', help='ajouter aussi les doublons probables')
    ap.add_argument('--sans-marque', action='store_true',
                    help='ne pas suffixer " (?)" les descriptions a confirmer')
    args = ap.parse_args()

    regles = charger_regles(args.regles)
    ops = lire_releve(args.releve)
    if args.depuis:
        d = datetime.fromisoformat(args.depuis)
        ops = [o for o in ops if o.date >= d]
    if args.jusqua:
        f = datetime.fromisoformat(args.jusqua) + timedelta(days=1)
        ops = [o for o in ops if o.date < f]

    classer(ops, regles)
    ops = agreger(ops)

    wb = openpyxl.load_workbook(args.budget)   # formules conservees
    ws = wb[FEUILLE]
    marquer_doublons(ops, lire_saisies(ws))
    if args.forcer:
        for o in ops:
            if o.statut == 'doublon?':
                o.statut = 'ajout'

    print(f'Releve : {args.releve}   ({len(ops)} operations retenues)')
    print(f'Budget : {args.budget}')
    rapport(ops, categories_valides(wb))

    ajouts = [o for o in ops if o.statut == 'ajout']
    if not args.ecrire:
        print('\n(simulation - rien n\'a ete ecrit ; ajouter --ecrire pour appliquer)')
        return
    if not ajouts:
        print('\nRien a ajouter.')
        return

    horodatage = datetime.now().strftime('%Y%m%d-%H%M%S')
    sauvegarde = args.budget.with_name(f'{args.budget.stem}.sauvegarde-{horodatage}.xlsx')
    shutil.copy2(args.budget, sauvegarde)

    depart = ecrire(ws, ajouts, marquer=not args.sans_marque)
    if etendre_recap(ws):
        print(f'Recap lateral : plages etendues jusqu\'a la ligne {FIN_VOULUE}.')
    wb.save(args.budget)

    print(f'\n{len(ajouts)} ligne(s) ecrite(s) a partir de la ligne {depart}.')
    print(f'Sauvegarde : {sauvegarde.name}')


if __name__ == '__main__':
    main()
